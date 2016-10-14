#!/usr/bin/env python
# coding=utf-8

# \_\_\_\_\_    \_\_\_                \_\_\_\_      \_\_\_
#     \_      \_      \_              \_      \_  \_      \_
#      \_      \_      \_  \_\_\_\_\_  \_      \_  \_      \_
#       \_      \_      \_              \_      \_  \_      \_
#        \_        \_\_\_                \_\_\_\_      \_\_\_

# [X] Separate sheets for sectors
# [X] Add easier way to input dates
# [X] Add separate files for dates and symbols
# [X] Figure out what to do about timestamps
# [X] Do said thing
# [X] Add weekday statistics at end
# [X] Check user dates against valid dates
# [ ] More feedback about dates (dots when date not found?)

# \_\_\_\_\_  \_      \_  \_\_\_\_      \_\_\_    \_\_\_\_    \_\_\_\_\_
#      \_      \_\_  \_\_  \_      \_  \_      \_  \_      \_      \_
#       \_      \_  \_  \_  \_\_\_\_    \_      \_  \_\_\_\_        \_
#        \_      \_      \_  \_          \_      \_  \_    \_        \_
#     \_\_\_\_\_  \_      \_  \_            \_\_\_    \_      \_      \_

import socket
import sys
import time
from datetime import datetime
from os.path import expanduser

import csv
import json
import openpyxl
import requests
# from lxml import html


# \_\_\_\_    \_\_\_\_\_  \_\_\_\_\_  \_\_\_\_\_  \_      \_  \_\_\_\_
#  \_      \_  \_          \_              \_      \_\_    \_  \_
#   \_      \_  \_\_\_\_\_  \_\_\_\_\_      \_      \_  \_  \_  \_\_\_\_
#    \_      \_  \_          \_              \_      \_    \_\_  \_
#     \_\_\_\_    \_\_\_\_\_  \_          \_\_\_\_\_  \_      \_  \_\_\_\_


def ifttt(action, v1='', v2='', v3=''):
    requests.post(
        'https://maker.ifttt.com/trigger/{0}/with/key/bgj70H05l-3HBccRCYvERV'
        .format(action), data={'value1': v1, 'value2': v2, 'value3': v3})


def get_sheet_corner(workbook_path, sheet_name=None):
    """Returns the column and row of the upper left corner of a spreadsheet.

    Indexing starts at 1, so A1 is (1, 1), not (0, 0)."""
    # I have to use x and y because rows and columns get me confused about
    # which way they go
    wb = openpyxl.load_workbook(workbook_path)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    first_x = 0
    corner_found = False
    while not corner_found:
        if first_x >= 1000:
            raise RuntimeError('No data found for 1000 columns')
        for x in range(first_x, -1, -1):
            y = first_x - x
            temp_cell = ws.cell(row=y + 1, column=x + 1)
            if temp_cell.value:
                return (x + 1, y + 1)
                corner_found = True
        first_x += 1


def read_sheet_column(workbook_path, sheet_name=None, headers=True):
    """Reads the first column in a given sheet.

    If headers is True, then loop through all the cells below the upper-left
    corner until a blank space is found. Return a list of all the cells. If a
    cell has a hashtag in the cell to the left of it, do not return that cell.
    """
    corner = get_sheet_corner(workbook_path, sheet_name)
    wb = openpyxl.load_workbook(workbook_path)
    output_cells = []
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    x = corner[0]
    if headers:
        y = corner[1] + 1  # Don't want the headers in the data
    else:
        y = corner[1]
    read_cell = ws.cell(row=y, column=x)
    while read_cell.value:
        read_cell = ws.cell(row=y, column=x)
        if x == 1:
            output_cells.append(read_cell.value)
        else:
            adjacent_cell = ws.cell(row=y-1, column=x)
            if '#' not in str(adjacent_cell.value):
                try:
                    output_cells.append(read_cell.value.upper())
                except AttributeError:
                    output_cells.append(read_cell.value)
        if output_cells[-1] is None:
            del output_cells[-1]
        y += 1
    return output_cells


def week(timestamp):
    """Returns the ISO calendar week number of a given timestamp.

    Timestamp can be either an integer or a string."""
    return datetime.utcfromtimestamp(int(timestamp)).isocalendar()[1]


def notify(message):
    """Gives a Pushbullet message."""
    ifttt('notify', v1=message)


def end_script(terminate=True):
    """Ends program."""
    if not isdev:
        input('Press enter to exit')
        sys.exit()
    elif terminate:
        sys.exit()


def error(msg):
    print(msg)
    notify(msg)
    end_script()


def excel_close(file):
    try:
        file.close()
        return True
    except PermissionError:
        return False
        input('Permissions denied! Please close all Excel windows and try'
              ' again.')
        if excel_close(file):
            pass


def rearrange(lst, order):
    """Returns lst but in the order of order.

    Indexing starts at 0."""
    return [lst[x] for x in order]


def mass_lookup(d, k):
    """Returns a list of the values of keys k from d."""
    output = []
    for key in k:
        try:
            output.append(d[key])
        except KeyError:
            output.append('')
    return output


# \_\_\_\_\_  \_\_\_\_\_  \_\_\_\_\_  \_\_\_\_    \_\_\_\_\_
#  \_              \_      \_      \_  \_      \_      \_
#   \_\_\_\_\_      \_      \_\_\_\_\_  \_\_\_\_        \_
#            \_      \_      \_      \_  \_    \_        \_
#     \_\_\_\_\_      \_      \_      \_  \_      \_      \_

version = '1.1.0'
print('Stock data downloader version {0}'.format(version))
isdev = socket.gethostname() == 'raphael'
if isdev:
    print('Developer mode active')
else:
    print('User mode active')

desktop = expanduser('~') + '\\Desktop\\'

print('Opening files...')
try:
    signs = read_sheet_column(desktop + 'stock_signs.xlsx')
except FileNotFoundError:
    write_signs = openpyxl.Workbook()
    write_signs.save(desktop + 'stock_signs.xlsx')
    print('Please go to your desktop and put the dates you want into'
          ' stock_signs.xlsx. Put hash marks in the cells to the left of the'
          ' ones you don\'t want.')
    end_script(terminate=False)
try:
    dates = read_sheet_column(desktop + 'stock_dates.xlsx')
    print(dates)
    dates_weeks = [date.isocalendar()[:1] for date in dates]
    print(dates_weeks)
except FileNotFoundError:
    write_dates = openpyxl.Workbook()
    write_dates.save(desktop + 'stock_dates.xlsx')
    print('Please go to your desktop and put the signs you want into'
          ' stock_dates.xlsx. Put hash marks in the cells to the left of the'
          " ones you don't want.")
    end_script(terminate=False)

backup_signs = [
    'A',
    'ABC',
    'ABT',
    'ACE',  # Renamed "ACET"
    'ACN',
    'ACT',  # Renamed "ATVI"
    'ADBE',
    'ADI',
    'AET',
    'AFL',
    'AGN',
    'AGU',
    'AIG',
    'ALL',
    'ALXN',
    'AMGN',
    'AMT',
    'AMZN',
    'APA',
    'APC',
    'APD',
    'AXP',
    'AZO',
    'BA',
    'BAC',
    'BAM',
    'BAX',
    'BBBY',
    'BDX',
    'BEN',
    'BFB',  # No industry/sector data
    'BHI',
    'BHP',
    'BIIB',
    'BMY',
    'BP',
    'BRK-B',
    'BUD',
    'BWA',
    'BXP',  # Idk
    'C',
    'CAH',
    'CAM',  # Renamed "CWH"
    'CAT',
    'CBS',
    'CELG',
    'CERN',
    'CHKP',
    'CI',
    'CMCSA',
    'CME',
    'CMG',
    'CMI',
    'CNQ',
    'COF',
    'COG',
    'COH',
    'COST',
    'COV',
    'CS',
    'CSCO',
    'CSX',
    'CTSH',
    'CTXS',
    'CVS',
    'CVX',
    'DAL',
    'DD',
    'DEO',
    'DFS',
    'DGX',
    'DHR',
    'DIS',
    'DLPH',
    'DOV',
    'DTV',  # No industry/sector data, no options data (valid stock tho)
    'DVA',
    'DVN',
    'EBAY',
    'ECL',
    'EL',
    'EMC',  # No industry/sector data
    'EMN',
    'ENB',
    'EOG',
    'EPD',
    'ESRX',
    'ESV',
    'ETN',
    'F',
    'FB',
    'FDX',
    'FIS',
    'FLR',
    'GD',
    'GE',
    'GILD',
    'GIS',
    'GLW',
    'GM',
    'GPS',
    'GSK',
    'GWW',
    'HAL',
    'HD',
    'HES',
    'HMC',
    'HOG',
    'HON',
    'HOT',  # No industry/sector data
    'HST',
    'HSY',
    'HUM',
    'ICE',
    'INTC',
    'IP',
    'ISRG',
    'JCI',
    'JNJ',
    'JPM',
    'KMP',  # Renamed ^KMP
    'KMX',
    'KO',
    'KR',
    'KRFT',  # No data...?
    'KSS',
    'L',
    'LLY',
    'LOW',
    'LVS',
    'LYB',
    'M',
    'MA',
    'MAR',
    'MAT',
    'MCD',
    'MCK',
    'MDLZ',
    'MDT',
    'MET',
    'MFC',
    'MHFI',  # Doesn't exist as a whole, "MHFInnn..." options do tho
    'MMC',
    'MO',
    'MON',
    'MOS',
    'MPC',
    'MRK',
    'MRO',
    'MRO',
    'MS',
    'MSFT',
    'MUR',
    'MYL',
    'NBL',
    'NE',
    'NEM',
    'NKE',
    'NLSN',
    'NOV',
    'NSC',
    'NUE',
    'NVS',  # Weirdness with one date
    'ORCL',
    'ORLY',
    'OXY',
    'PCP',  # Renamed "^PCP"
    'PEP',
    'PFE',
    'PG',
    'PH',
    'PM',
    'PNC',
    'PNR',
    'PPG',
    'PRU',
    'PSX',
    'PX',
    'PXD',
    'QCOM',
    'QQQ',  # No industry/sector data
    'REGN',
    'RIO',
    'RL',
    'ROP',
    'ROST',
    'RRC',
    'RSG',
    'SBUX',
    'SE',
    'SHW',
    'SJM',
    'SLB',
    'SLM',
    'SNDK',  # Exists only in options form
    'SPG',
    'STT',
    'STZ',
    'SU',
    'SWK',
    'SYK',
    'TCK',
    'TEL',
    'TJX',
    'TM',
    'TMO',
    'TROW',
    'TRV',
    'TWC',  # ^TWC
    'TWX',
    'TYC',  # No industry/sector data
    'UAL',
    'UNH',
    'UNP',
    'UPS',
    'UTX',
    'V',
    'VFC',
    'VIAB',
    'VLO',
    'VNO',
    'VZ',
    'WAG',  # ^WAG
    'WDC',
    'WFC',
    'WFM',
    'WMB',
    'WMT',
    'WY',
    'WYNN',
    'YHOO',
    'YUM',
    'ZMH'  # ^ZMH
]

signs = backup_signs  # Until I can get a list of better ones from Grandpa

assert signs
assert dates

print('{0} signs, {1} dates'.format(len(signs), len(dates)))

dt = datetime.fromtimestamp(time.time())
date = dt.strftime('%d-%m-%Y')

print(date)

if not isdev:
    output_path = (
        'C:/Users/Gary/Documents/Option_tables/Option_Model_Files/'
        'OptionReportDirectory/options_report_{0}.csv'.format(date)
    )
else:
    output_path = 'options_report_{0}.csv'.format(date)

output_name = output_path.split('/')[-1]

start = time.time()

# \_\_\_\_      \_\_\_    \_      \_  \_      \_
#  \_      \_  \_      \_  \_      \_  \_\_    \_
#   \_      \_  \_      \_  \_  \_  \_  \_  \_  \_  \_\_\_\_\_
#    \_      \_  \_      \_  \_  \_  \_  \_    \_\_
#     \_\_\_\_      \_\_\_      \_  \_    \_      \_

#       \_            \_\_\_      \_\_\_    \_\_\_\_
#        \_          \_      \_  \_      \_  \_      \_
#         \_          \_      \_  \_\_\_\_\_  \_      \_
#          \_          \_      \_  \_      \_  \_      \_
#           \_\_\_\_\_    \_\_\_    \_      \_  \_\_\_\_

options_data_url = 'https://query1.finance.yahoo.com/v7/finance/options/{0}'
stock_data_url = ('https://query1.finance.yahoo.com/v10/finance/quoteSummary/'
                  '{0}?modules=assetProfile')

all_data = [[  # Headers
    'Stock', 'Timestamp', 'Contract Symbol', 'Strike', 'Currency',
    'Last Price', 'Change', '% Change', 'Volume', 'Open Interest', 'Bid',
    'Ask', 'Contract Size', 'Expiration', 'Last Trade Date',
    'Implied Volatility', 'In The Money', 'Stock Last', 'Industry', 'Sector',
    'Company']]
json_headers = [
    'contractSymbol', 'strike', 'currency', 'lastPrice',
    'change', 'percentChange', 'volume', 'openInterest', 'bid', 'ask',
    'contractSize', 'expiration', 'lastTradeDate', 'impliedVolatility',
    'inTheMoney', 'quoteLast', 'industry', 'sector', 'company']
errors = []

for sign in signs:
    # all_data[sign] = {}
    print('\n{0:{1}} ({2:{3}} of {4})'.format(sign,
                                              len(max(signs, key=len)),
                                              signs.index(sign) + 1,
                                              len(str(len(signs))),
                                              len(signs)),
          end='')

    dates_page = requests.get(options_data_url.format(sign))
    dates_json = json.loads(dates_page.text)
    try:
        timestamps_from_site = (
            dates_json['optionChain']['result'][0]['expirationDates']
        )
    except (IndexError, TypeError) as e:
        print(' Non-existent', end='')
        continue

    # timestamps_to_use = timestamps_from_site
    timestamps_to_use = [ts for ts in timestamps_from_site if datetime.fromtimestamp(ts).isocalendar()[:1] in dates_weeks]
    print(' [', '-' * len(timestamps_to_use), ']', sep='', end='', flush=True)

    weekdays = []

    stock_page = requests.get(stock_data_url.format(sign))
    stock_json = json.loads(stock_page.text)

    profile = stock_json['quoteSummary']['result'][0]['assetProfile']
    try:
        industry, sector = profile['industry'], profile['sector']
    except KeyError:
        print(' Sector and industry unavailable', '\b' * 32,
              sep='', end='', flush=True)
        industry = sector = ''

    print('\b' * (len(timestamps_to_use) + 1), end='')

    messages = []
    for ts in timestamps_to_use:
        complete_success = True
        try:
            data_page = requests.get(
                options_data_url.format(sign) + '?date=' + str(ts)
            )
        except TimeoutError:  # Redundancy just to be sure
            try:
                data_page = requests.get(
                    options_data_url.format(sign) + '?date=' + str(ts)
                )
            except TimeoutError:
                # TODO: More verbose
                print('-', end='', flush=True)
                messages.append('{d} timed out'.format(
                    d=datetime.utcfromtimestamp(ts).strftime('%m/%d/%Y')
                ))
                continue
        data_json = json.loads(data_page.text)
        specific_data = data_json['optionChain']['result'][0]
        # {'Stock Last': specific_data['quote']['regularMarketPrice']}
        data_dict = (specific_data['options'][0]['calls'])  # List of dicts
        for row in data_dict:
            row.update(
                {'quoteLast': specific_data['quote']['regularMarketPrice'],
                 'company': specific_data['quote']['longName'],
                 'industry': industry,
                 'sector': sector}
            )
            try:
                all_data.append([sign, start] + [row[key]
                                for key in json_headers])
            except KeyError:
                # TODO: More verbose
                complete_success = False
                messages.append('Something went wrong with {d} ({ts})'.format(
                    d=datetime.utcfromtimestamp(ts).strftime('%m/%d/%Y'),
                    ts=ts
                ))
                print('-', end='', flush=True)
                continue
        if complete_success:
            print('=', end='', flush=True)
    if messages:
        print('] ', ', '.join(messages), end='', flush=True)

print()  # Allow printing of the last line

# \_\_\_\_\_    \_\_\_    \_\_\_\_    \_      \_  \_\_\_\_\_  \_\_\_\_\_
#  \_          \_      \_  \_      \_  \_\_  \_\_  \_      \_      \_
#   \_\_\_\_    \_      \_  \_\_\_\_    \_  \_  \_  \_\_\_\_\_      \_
#    \_          \_      \_  \_    \_    \_      \_  \_      \_      \_
#     \_            \_\_\_    \_      \_  \_      \_  \_      \_      \_

headers = [
    'Symbol',
    'Company',
    'Industry',
    'Sector',
    'Price',
    'Expiration',
    'Strike',
    'Bid',
    'Ask',
    'Volume',
    'Last Call',
    datetime.now().date(),
    'days',
    '70,000',
    ' $invested',
    '$prem',
    'prem%',
    'annPrem%',
    'MaxRet',
    'Max%',
    'annMax%',
    '10%'
]

all_data_by_header = [
    {h: x[i] for i, h in enumerate(all_data[0])} for x in all_data[1:]
]

for d in all_data_by_header:
    d['Timestamp'] = (
        datetime.utcfromtimestamp(d['Timestamp']).strftime('%m/%d/%Y %H:%M'))

    d['Last Trade Date'] = (
        datetime.utcfromtimestamp(d['Last Trade Date']).strftime('%m/%d/%Y'))

    d['Expiration'] = (
        datetime.utcfromtimestamp(d['Expiration']).strftime('%m/%d/%Y'))

# Original formulas
# formulas = [
#     '=IF(H{n}<F{n},(H{n}-F{n})+I{n},I{n})',
#     '=G{n}-M$6',
#     '=ROUND(O$6/((F{n}-0)*100),0)',
#     '=100*O{n}*(F{n}-0)',
#     '=100*M{n}*O{n}',
#     '=(Q{n}/P{n})*100',
#     '=(365/N{n})*R{n}',
#     '=100*M{n}*O{n}',
#     '=(T{n}/P{n})*100',
#     '=((365/N{n})*U{n})*100',
#     '=IF((ABS(H{n}-F{n})/H{n})<W$6,"NTM","")'
# ]

formulas = [
    '=IF(K{n}<I{n},(K{n}-I{n})+O{n},O{n})',
    '=J{n}-P$6',
    '=ROUND(R$6/((I{n}-0)*100),0)',
    '=100*R{n}*(I{n}-0)',
    '=100*P{n}*R{n}',
    '=T{n}/S{n}',
    '=(365/Q{n})*U{n}',
    '=IF(K{n}>I{n},(100*R{n}*(K{n}-I{n}))+T{n},T{n})',
    '=W{n}/S{n}',
    '=(365/Q{n})*X{n}',
    '=IF((ABS(K{n}-I{n})/K{n})<Z$6,"NTM","")'
]

v_offset = 5
h_offset = 4

formatted_data_table = [headers] + [
    mass_lookup(row, [
        'Stock', 'Company', 'Industry', 'Sector', 'Stock Last', 'Expiration',
        'Strike', 'Bid', 'Ask', 'Volume', 'Last Price']) +
    # +2 because Excel starts counting at 1 and because there are headers
    [f.format(n=i+v_offset+2) for f in formulas]
    for i, row in enumerate(all_data_by_header)]

# Offset for formulas to work
formatted_data_table = (
    [[]] * v_offset + [[''] * h_offset + row for row in formatted_data_table]
)

# # \_      \_  \_\_\_\_    \_\_\_\_\_  \_\_\_\_\_  \_\_\_\_\_
# #  \_      \_  \_      \_      \_          \_      \_
# #   \_      \_  \_\_\_\_        \_          \_      \_\_\_\_\_
# #    \_  \_  \_  \_    \_        \_          \_      \_
# #     \_\_  \_\_  \_      \_  \_\_\_\_\_      \_      \_\_\_\_\_

print('Writing to {0}...'.format(output_name), end=' ', flush=True)

try:
    with open(output_path, 'w', newline='') as csv_file:
        csv_writer = csv.writer(csv_file)
        for row in formatted_data_table:
            csv_writer.writerow(row)
except PermissionError:
    print('\a\nPlease close {0}.'.format(output_name))
    input('Press enter when done')
    print('Writing to {0}...'.format(output_name, end=' ', flush=True))
    with open(output_path, 'w', newline='') as csv_file:
        csv_writer = csv.writer(csv_file)
        for row in formatted_data_table:
            csv_writer.writerow(row)

print('Done')
